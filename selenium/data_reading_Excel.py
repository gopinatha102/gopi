from selenium import webdriver
import openpyxl
path=r"C:\Users\PAVAN KUMARA\Downloads\manual text.xlsx"


workbook=openpyxl.load_workbook(path)
sheet=workbook.active
#sheet=workbook.get_sheet_by_name("Test_case")
rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end=" ")
    print()